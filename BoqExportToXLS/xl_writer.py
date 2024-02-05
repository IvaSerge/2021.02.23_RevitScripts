import clr
import os
import sys
import shutil

local_data = os.getenv("LOCALAPPDATA")
dyn_path = r"\python-3.9.12-embed-amd64\Lib"
py_path = local_data + dyn_path
sys.path.append(py_path)


# ================ Revit imports
clr.AddReference('RevitAPI')
from Autodesk.Revit.DB import *

clr.AddReference('RevitAPIUI')
from Autodesk.Revit.UI import *

clr.AddReference("RevitServices")
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager

# ================ Python imports
from System import Array
from System.Collections.Generic import *
from importlib import reload
import openpyxl

from PIL import Image

# ================ local imports
import toolsrvt
reload(toolsrvt)
from toolsrvt import *
import db_reader
reload(db_reader)
from db_reader import *


def create_files_names(boq_name, rev_doc_number, boq_descr, path_to_save):
	name_number = boq_name
	name_prefix = "_XLSX"
	name_rev = f'[{int(rev_doc_number):02d}]'
	name_description = boq_descr

	# xlsx file name
	name_xlsx = boq_name
	name_xlsx += name_prefix + name_rev
	if "BOQ" in name_description:
		name_xlsx += " - " + name_description
	else:
		name_xlsx += " - BOQ_" + name_description
	name_xlsx += ".xlsx"

	# pdf file name
	name_pdf = name_number
	name_pdf += name_rev
	if "BOQ" in name_description:
		name_pdf += " - " + name_description
	else:
		name_pdf += " - BOQ_" + name_description
	name_pdf += ".pdf"
	
	# check file names
	check_file_name(name_xlsx)
	check_file_name(name_pdf)

	# add path to beginning
	name_xlsx = path_to_save + "\\" + name_xlsx
	name_pdf = path_to_save + "\\" + name_pdf

	return name_xlsx, name_pdf

def move_template_xls_file(dyn_path, xl_save_to):
	template_path = dyn_path + "\\boq_template.xlsx"
	shutil.copy(template_path, xl_save_to)

def write_first_page(path_xlsx, boq_name, rev_doc_number, seq_number, doc):

	# get project info
	proj_status = doc.ProjectInformation.Status
	proj_address = doc.ProjectInformation.Address

	# project Discipline is project specific
	proj_discipline = doc.ProjectInformation.LookupParameter("Discipline")
	if proj_discipline:
		proj_discipline = doc.ProjectInformation.LookupParameter("Discipline").AsString()
	elif doc.ProjectInformation.LookupParameter("TSLA_ProjectDiscipline"):
		proj_discipline = doc.ProjectInformation.LookupParameter("TSLA_ProjectDiscipline").AsString()
	else:
		raise ValueError("Project Dyscipline not found")

	proj_str = proj_status + ". "
	proj_str += proj_address + ". "
	proj_str += proj_discipline

	# get revision info
	revisions = FilteredElementCollector(doc).\
		OfClass(Autodesk.Revit.DB.Revision).ToElements()
	revision = [i for i in revisions if i.SequenceNumber == seq_number]

	if not revision:
		raise ValueError("Revision not found")
	else:
		revision = revision[0]

	rev_date = revision.RevisionDate
	rev_description = revision.Description
	rev_issued_by = revision.IssuedBy

	# Cover page - write info
	wb = openpyxl.load_workbook(path_xlsx)
	wb.active = wb["Cover"]
	ws = wb.active

	ws["A32"] = proj_str
	ws["A34"] = boq_name
	ws["A44"] = rev_date
	ws["B44"] = rev_doc_number
	ws["C44"] = rev_description
	ws["H44"] = rev_issued_by

	# check cell height based on revision description
	# max str length is 47
	rev_description_height = math.ceil(
		len(rev_description) / 47)
	ws.row_dimensions[44].height = 15.75 * rev_description_height

	# save to new path
	wb.save(path_xlsx)


def write_totals(xl_path, boq_list):

	# Second page - write info
	wb = openpyxl.load_workbook(xl_path)
	wb.active = wb["BOQ Totals"]
	ws: openpyxl.Workbook.worksheets = wb.active

	# set columns width
	ws.column_dimensions["A"].width = 50
	ws.column_dimensions["B"].width = 10
	ws.column_dimensions["C"].width = 45
	ws.column_dimensions["D"].width = 30

	rw = 1
	for boq_object in boq_list:
		rw += 1
		boq_text_shedule = boq_object.boq
		for r_count, row in enumerate(boq_text_shedule, 1):
			for clmn, val in enumerate(row, 1):
				current_cell = ws.cell(row=rw, column=clmn)

				# get and set cell style by row
				if r_count == 1:
					# first line
					style_cell = "DiRootsFullNameTitleStyle"

				elif r_count == 2:
					# second line
					style_cell = "DiRootsHeaderStyle"

				else:
					# other lines
					style_cell = "Normal"

				current_cell.style = style_cell
				if val:
					current_cell.value = val
			rw += 1

	# set print area
	first_cell = "A1"
	last_cell = ws.cell(row=rw, column=4).coordinate
	ws.print_area = first_cell + ":" + last_cell

	wb.save(xl_path)
